"""
Parser dla PracownicyAPT_list_ver_02_NEW.xlsx.

Struktura arkusza PracownicyAPT01 (indeksy kolumn 0-based):
  0:  nazwisko
  1:  imię
  2–6: oceny 1–5
  7:  nazwa_agencji
  8–10: oceny 6–8
  11: płeć
  12: grupa
  13–18: oceny 9–14

Wartości ocen w pliku testowym są puste — plik jest szablonem.
"""
import openpyxl

# Kolumna → numer oceny (1-14)
SCORE_COLS: dict[int, int] = {
    2: 1, 3: 2, 4: 3, 5: 4, 6: 5,
    8: 6, 9: 7, 10: 8,
    13: 9, 14: 10, 15: 11, 16: 12, 17: 13, 18: 14,
}


def parsuj_pracownikow_apt(plik) -> tuple[list[dict], list[str]]:
    """
    Zwraca (pracownicy_list, ostrzezenia).
    Każdy pracownik: {'nazwisko', 'imie', 'nazwa_agencji', 'plec', 'grupa', 'oceny': {numer_str: wartosc}}
    """
    wb = openpyxl.load_workbook(plik, data_only=True)

    ws = wb['PracownicyAPT01'] if 'PracownicyAPT01' in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    pracownicy: list[dict] = []
    ostrzezenia: list[str] = []

    if not rows:
        return [], ['Pusty arkusz']

    for row in rows[1:]:  # pomiń wiersz nagłówkowy
        if not row:
            continue
        nazwisko_raw = row[0] if len(row) > 0 else None
        imie_raw = row[1] if len(row) > 1 else None
        if not nazwisko_raw and not imie_raw:
            continue

        nazwisko = str(nazwisko_raw).strip() if nazwisko_raw else ''
        imie = str(imie_raw).strip() if imie_raw else ''
        if not nazwisko:
            continue

        agencja_raw = row[7] if len(row) > 7 else None
        plec_raw = row[11] if len(row) > 11 else None
        grupa_raw = row[12] if len(row) > 12 else None

        oceny: dict[str, float | None] = {}
        for ci, numer in SCORE_COLS.items():
            if ci < len(row) and row[ci] is not None:
                try:
                    oceny[str(numer)] = float(row[ci])
                except (TypeError, ValueError):
                    pass

        pracownicy.append({
            'nazwisko': nazwisko,
            'imie': imie,
            'nazwa_agencji': str(agencja_raw).strip() if agencja_raw else '',
            'plec': str(plec_raw).strip() if plec_raw else '',
            'grupa': str(grupa_raw).strip() if grupa_raw else '',
            'oceny': oceny,
        })

    if not pracownicy:
        ostrzezenia.append('Nie znaleziono żadnych wierszy z danymi pracowników APT.')

    return pracownicy, ostrzezenia
