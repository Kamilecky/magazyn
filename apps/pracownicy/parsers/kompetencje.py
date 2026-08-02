"""
Parser dla KOMPETENCJE_PRACOWNIKÓW_ACT_NEW.xlsx.

Struktura:
- Arkusz "Wynik finalny" (lub wiele arkuszy — oba warianty obsługiwane)
- Wiersze 3–5: 3-poziomowy nagłówek (Dział / Poddział / Aktywność), scalone komórki → forward-fill
- Wiersz 6:   etykiety kolumn pracownika (A–N) i nagłówki kompetencji
- Wiersz 7+:  dane pracowników; kolumny O+ = wynik kompetencji (0–50)

Kolumny A–N (0–13) mapowane na pola Pracownik:
  0: departament, 1: nazwisko, 2: imię, 3: prefix (pomijany),
  4: Nazwisko i Imię (pomijany), 5: nr_ewidencyjny, 6: data_zatrudnienia,
  7: stanowisko, 8: strefa, 9: dział, 10: zmiana, 11: zmiana_grupa,
  12: przełożony, 13: komentarz
"""
from datetime import date, datetime
import openpyxl

# Etykiety działów Prasa w różnych wariantach pisowni — kolumny z prasa są pomijane
PRASA_ETYKIETY = {'prasa', 'PRASA', 'Prasa'}

# Pierwsze 14 kolumn (A–N) to dane pracownika; od kolumny O zaczynają się wyniki kompetencji
PIERWSZE_KOLUMNY = 14


# Wypełnia brakujące wartości w scalonej komórce — kopiuje ostatnią niepustą wartość w prawo
# Potrzebne bo openpyxl zwraca None dla scalonych komórek nie będących komórką główną
def _forward_fill(row: tuple) -> list:
    result = list(row)
    last = None
    for i, v in enumerate(result):
        s = str(v).strip() if v is not None else ''
        if s:
            last = s
            result[i] = s
        else:
            result[i] = last
    return result


# Bezpieczna konwersja wartości komórki na string; None → pusty string
def _safe_str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


# Parsuje datę z różnych formatów (datetime z Excel, ISO, polskie dd.mm.yyyy)
def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.isoformat()[:10]
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def parsuj_kompetencje(plik) -> tuple[list[dict], dict, list[str]]:
    """
    Zwraca (pracownicy_list, kompetencje_dict, ostrzezenia).
      pracownicy_list: [{'nazwisko': ..., 'imie': ..., 'dzial': ..., ...}]
      kompetencje_dict: {(nazwisko, imie): [{'aktywnosc_nazwa': ..., 'aktywnosc_dzial': ..., 'wynik': ...}]}
    """
    wb = openpyxl.load_workbook(plik, data_only=True)

    # Preferuje arkusz "Wynik finalny"; fallback na wszystkie arkusze (stary format pliku)
    if 'Wynik finalny' in wb.sheetnames:
        sheets = [wb['Wynik finalny']]
    else:
        sheets = [wb[n] for n in wb.sheetnames]

    all_pracownicy: dict[tuple, dict] = {}
    all_kompetencje: dict[tuple, list] = {}
    ostrzezenia: list[str] = []

    # Parsuj każdy arkusz osobno i scal wyniki (pracownik z wielu arkuszy → jeden rekord)
    for ws in sheets:
        p, k, o = _parsuj_arkusz(ws)
        for key, data in p.items():
            if key not in all_pracownicy:
                all_pracownicy[key] = data
        for key, komp in k.items():
            all_kompetencje.setdefault(key, []).extend(komp)
        ostrzezenia.extend(o)

    return list(all_pracownicy.values()), all_kompetencje, ostrzezenia


def _parsuj_arkusz(ws) -> tuple[dict, dict, list]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 7:
        return {}, {}, ['Arkusz za krótki (< 7 wierszy)']

    # Wiersze 3–5 (0-indexed: 2–4) to trzypoziomowy nagłówek: Dział / Poddział / Aktywność
    # forward_fill wypełnia scalone komórki (Excel scalone = None poza pierwszą komórką)
    row3 = _forward_fill(rows[2])   # Dział
    row4 = _forward_fill(rows[3])   # Poddział
    row5 = list(rows[4])            # Aktywność — zazwyczaj nie scalane

    # Zbuduj mapę: numer_kolumny → {dzial, aktywnosc} dla kolumn kompetencji (od 14)
    col_map: dict[int, dict] = {}
    for i in range(PIERWSZE_KOLUMNY, max(len(row3), len(row5))):
        dzial = row3[i] if i < len(row3) else None
        akt_raw = row5[i] if i < len(row5) else None
        if not akt_raw:
            continue
        akt = str(akt_raw).strip()
        if not akt:
            continue
        # Pomijamy kolumny z działu Prasa (nie wchodzą do macierzy kompetencji)
        if dzial and dzial.lower() == 'prasa':
            continue
        col_map[i] = {
            'dzial': str(dzial or '').strip(),
            'aktywnosc': akt,
        }

    pracownicy: dict[tuple, dict] = {}
    kompetencje: dict[tuple, list] = {}

    # Dane pracowników zaczynają się od wiersza 7 (0-indexed: 6)
    for row in rows[6:]:
        if not row:
            continue
        nazwisko_raw = row[1] if len(row) > 1 else None
        imie_raw = row[2] if len(row) > 2 else None
        # Pomiń wiersze bez nazwiska/imienia (wiersze sumaryczne, puste)
        if not nazwisko_raw or not imie_raw:
            continue
        nazwisko = _safe_str(nazwisko_raw)
        imie = _safe_str(imie_raw)
        if not nazwisko or not imie:
            continue

        # Klucz deduplicacji: (nazwisko, imie) — ten sam pracownik w różnych arkuszach
        key = (nazwisko, imie)
        if key not in pracownicy:
            pracownicy[key] = {
                'nazwisko': nazwisko,
                'imie': imie,
                'departament': _safe_str(row[0] if len(row) > 0 else None),
                'nr_ewidencyjny': _safe_str(row[5] if len(row) > 5 else None) or None,
                'data_zatrudnienia': _parse_date(row[6] if len(row) > 6 else None),
                'stanowisko': _safe_str(row[7] if len(row) > 7 else None),
                'strefa': _safe_str(row[8] if len(row) > 8 else None),
                'dzial': _safe_str(row[9] if len(row) > 9 else None),
                'zmiana': _safe_str(row[10] if len(row) > 10 else None),
                'zmiana_grupa': _safe_str(row[11] if len(row) > 11 else None),
                'przelozony': _safe_str(row[12] if len(row) > 12 else None),
                'komentarz': _safe_str(row[13] if len(row) > 13 else None),
            }
            kompetencje[key] = []

        # Odczytaj wyniki kompetencji dla każdej zmapowanej kolumny
        for ci, col_info in col_map.items():
            if ci >= len(row):
                continue
            val = row[ci]
            if val is None:
                continue
            try:
                wynik = float(val)
            except (TypeError, ValueError):
                continue
            # Zapisujemy tylko wyniki > 0 (brak kompetencji = brak wiersza w DB)
            if wynik > 0:
                kompetencje[key].append({
                    'aktywnosc_nazwa': col_info['aktywnosc'],
                    'aktywnosc_dzial': col_info['dzial'],
                    'wynik': wynik,
                })

    return pracownicy, kompetencje, []
