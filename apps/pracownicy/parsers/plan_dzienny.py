"""
Parser dla Plan_dzienny_NEW.xlsx.

Struktura pliku:
- Jeden arkusz "PLAN NEW"
- Sekcje działów wykrywane przez kolumna B == 'Bufor'
- Każdy wiersz aktywności zawiera 24 punkty godzinowe (3 zmiany × 8 godzin)

Indeksy kolumn (0-based):
  Zmiana I:   kolumny 11–18 (L–S),   godziny 6, 7, 8, 9, 10, 11, 12, 13
  Zmiana II:  kolumny 22–29 (W–AD),  godziny 14, 15, 16, 17, 18, 19, 20, 21
  Zmiana III: kolumny 33–40 (AH–AO), godziny 22, 23, 0, 1, 2, 3, 4, 5
"""
from dataclasses import dataclass, field
from typing import Optional
import openpyxl

# Mapowanie numeru zmiany na listę godzin doby (zmiana III obejmuje północ)
ZMIANA_GODZINY = {
    1: [6, 7, 8, 9, 10, 11, 12, 13],
    2: [14, 15, 16, 17, 18, 19, 20, 21],
    3: [22, 23, 0, 1, 2, 3, 4, 5],
}

# Indeksy kolumn Excel (0-based) dla wolumenu i godzin każdej zmiany
COL_WOLUMEN_I = 9    # J
COL_HOURS_I = 11     # L
COL_WOLUMEN_II = 20  # U
COL_HOURS_II = 22    # W
COL_WOLUMEN_III = 31 # AF
COL_HOURS_III = 33   # AH


# Struktura danych dla jednego wiersza planu (jedna aktywność ze wszystkimi zmianami)
@dataclass
class WierszPlanu:
    dzial: str
    aktywnosc: str
    wolumen_I: float
    wolumen_II: float
    wolumen_III: float
    godziny: dict  # {zmiana_int: {godzina_int: liczba_osob_float}}


# Bezpieczna konwersja wartości komórki na float — obsługuje None, int, float, str z błędami formuł
def _to_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        # Pusta komórka lub błąd formuły Excel (#REF!, #DIV/0! itp.) → 0
        if not s or s.startswith('#'):
            return 0.0
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return 0.0
    return 0.0


# Sprawdza czy wartość komórki to błąd formuły Excel (np. #REF!, #VALUE!)
def _is_error(val) -> bool:
    return isinstance(val, str) and val.strip().startswith('#')


def parsuj_plan_dzienny(plik) -> tuple[list[WierszPlanu], list[str]]:
    """
    Zwraca (wiersze, ostrzezenia).
    Każdy WierszPlanu zawiera 24 rekordy godzinowe w polu `godziny`.
    """
    # data_only=True — odczytuje wartości zamiast formuł Excel
    wb = openpyxl.load_workbook(plik, data_only=True)

    # Preferuje arkusz o nazwie "PLAN NEW"; fallback na aktywny arkusz
    ws = wb['PLAN NEW'] if 'PLAN NEW' in wb.sheetnames else wb.active

    wiersze: list[WierszPlanu] = []
    ostrzezenia: list[str] = []
    aktualny_dzial: Optional[str] = None

    for row in ws.iter_rows(values_only=True):
        # Pomijaj puste wiersze
        if not row or all(c is None for c in row):
            continue

        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None

        # Wiersz nagłówka sekcji: kolumna B == 'Bufor' → kolumna A to nazwa działu
        if isinstance(col_b, str) and col_b.strip() == 'Bufor':
            if col_a is not None:
                aktualny_dzial = str(col_a).strip()
            continue

        # Pomiń wiersze przed pierwszym nagłówkiem działu
        if aktualny_dzial is None or not col_a:
            continue

        aktywnosc = str(col_a).strip()
        # Pomiń wiersz sumaryczny działu (gdy nazwa aktywności == nazwa działu)
        if not aktywnosc or aktywnosc == aktualny_dzial:
            continue

        # Zbierz ostrzeżenia o błędach formuł w kolumnach wolumenu
        for col_i in (COL_WOLUMEN_I, COL_WOLUMEN_II, COL_WOLUMEN_III):
            if col_i < len(row) and _is_error(row[col_i]):
                ostrzezenia.append(
                    f'Błąd formuły {row[col_i]} — wiersz „{aktywnosc}" '
                    f'(dział {aktualny_dzial}), traktowany jako 0'
                )

        wolumen_I = _to_float(row[COL_WOLUMEN_I] if COL_WOLUMEN_I < len(row) else None)
        wolumen_II = _to_float(row[COL_WOLUMEN_II] if COL_WOLUMEN_II < len(row) else None)
        wolumen_III = _to_float(row[COL_WOLUMEN_III] if COL_WOLUMEN_III < len(row) else None)

        # Odczytaj 8 wartości godzinowych dla każdej z 3 zmian (razem 24 punkty)
        godziny = {}
        for zmiana, start_col, hours in (
            (1, COL_HOURS_I, ZMIANA_GODZINY[1]),
            (2, COL_HOURS_II, ZMIANA_GODZINY[2]),
            (3, COL_HOURS_III, ZMIANA_GODZINY[3]),
        ):
            g = {}
            for i, h in enumerate(hours):
                ci = start_col + i
                g[h] = _to_float(row[ci] if ci < len(row) else None)
            godziny[zmiana] = g

        wiersze.append(WierszPlanu(
            dzial=aktualny_dzial,
            aktywnosc=aktywnosc,
            wolumen_I=wolumen_I,
            wolumen_II=wolumen_II,
            wolumen_III=wolumen_III,
            godziny=godziny,
        ))

    return wiersze, ostrzezenia
