"""
Parser dla Struktura___Grafik___Absencje_NEW.xlsx.

Struktura:
- Arkusze: "Struktura IB", "Struktura OB", "Struktura FF", "Struktura PR", "Struktura ZW"
- Arkusz "Listy Rozwijane": źródło prawdy dla typów absencji (kolumna B)
- Wiersze 1–5: statystyki zbiorcze (pomijane)
- Wiersz 6: nagłówki kolumn
- Wiersz 7+: dane pracowników; kolumny z datami = obecność/absencja
"""
from datetime import date, datetime
import openpyxl

# Mapowanie nazwy arkusza na dział — obsługuje warianty nazw (IB/IN oba = Inbound)
ARKUSZE_DZIALY = {
    'Struktura IN': 'Inbound',
    'Struktura IB': 'Inbound',
    'Struktura OB': 'Outbound',
    'Struktura FF': 'Fulfilment',
    'Struktura PR': 'Prasa',
    'Struktura ZW': 'Zwroty',
}

# Nagłówki kolumn w pliku → nazwy pól modelu Pracownik
PRACOWNIK_NAGLOWKI = {
    'Nazwisko': 'nazwisko',
    'Imię': 'imie',
    'Nr ewidencyjny': 'nr_ewidencyjny',
    'Data zatrudnienia': 'data_zatrudnienia',
    'Stanowisko': 'stanowisko',
    'Strefa': 'strefa',
    'Dział': 'dzial',
    'Zmiana': 'zmiana',
    'Zmiana grupa': 'zmiana_grupa',
    'Przełożony': 'przelozony',
}

# Znormalizowane klucze (lowercase, bez spacji/podkreśleń) → pole modelu
# Normalizacja pozwala dopasować nagłówki niezależnie od spacji i wielkości liter
_NAGLOWKI_NORM = {
    k.lower().replace(' ', '').replace('_', ''): v
    for k, v in PRACOWNIK_NAGLOWKI.items()
}

# Domyślny typ absencji jeśli arkusz "Listy Rozwijane" nie istnieje
DOMYSLNE_TYPY_ABSENCJI = {'Nieobecny'}


def _safe_str(val) -> str:
    return str(val).strip() if val is not None else ''


# Parsuje datę z Excel (datetime obiektu) lub różnych formatów stringowych
def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.isoformat()[:10]
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


# Próbuje sparsować datę — zwraca None jeśli się nie uda (używane dla nagłówków kolumn dat)
def _try_parse_date(val) -> str | None:
    if isinstance(val, (date, datetime)):
        return val.isoformat()[:10]
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y'):
            try:
                return datetime.strptime(val.strip(), fmt).date().isoformat()
            except ValueError:
                pass
    return None


def parsuj_strukture(plik) -> tuple[list[dict], list[dict], list[str]]:
    """
    Zwraca (pracownicy_list, absencje_list, ostrzezenia).
    """
    wb = openpyxl.load_workbook(plik, data_only=True)

    # Wczytaj listę typów absencji z arkusza pomocniczego (kolumna B od wiersza 2)
    typy_absencji = set(DOMYSLNE_TYPY_ABSENCJI)
    if 'Listy Rozwijane' in wb.sheetnames:
        ws_lr = wb['Listy Rozwijane']
        for row in ws_lr.iter_rows(values_only=True, min_row=2):
            val = row[1] if len(row) > 1 else None
            if val and isinstance(val, str) and val.strip():
                typy_absencji.add(val.strip())

    all_pracownicy: dict[tuple, dict] = {}
    all_absencje: list[dict] = []
    ostrzezenia: list[str] = []

    # Przetwarzaj tylko arkusze ze strukturą (IB/OB/FF/PR/ZW)
    dostepne_arkusze = [n for n in wb.sheetnames if n in ARKUSZE_DZIALY]
    if not dostepne_arkusze:
        ostrzezenia.append(
            'Nie znaleziono żadnego arkusza struktury '
            '(oczekiwano: Struktura IB / OB / FF / PR / ZW).'
        )
        return [], [], ostrzezenia

    for sheet_name in dostepne_arkusze:
        ws = wb[sheet_name]
        p, a, o = _parsuj_arkusz(ws, sheet_name, typy_absencji)
        # Dane z późniejszych arkuszy nadpisują wcześniejsze (ten sam pracownik w wielu działach)
        for key, data in p.items():
            all_pracownicy[key] = {**all_pracownicy.get(key, {}), **data}
        all_absencje.extend(a)
        ostrzezenia.extend(o)

    return list(all_pracownicy.values()), all_absencje, ostrzezenia


def _parsuj_arkusz(ws, sheet_name: str, typy_absencji: set) -> tuple[dict, list, list]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 7:
        return {}, [], [f'Arkusz {sheet_name}: zbyt mało wierszy']

    # Wiersz 6 (0-indexed: 5) = nagłówki kolumn pracownika i kolumny dat
    header_row = rows[5]

    col_indices: dict[str, int] = {}   # nazwa_pola → indeks kolumny
    date_cols: dict[int, str] = {}     # indeks_kolumny → data ISO (kolumny z datami absencji)

    for i, header in enumerate(header_row):
        if header is None:
            continue
        header_str = str(header).strip()
        # Normalizacja: usuń spacje i podkreślenia, małe litery — odporna na literówki w nagłówkach
        header_norm = header_str.lower().replace(' ', '').replace('_', '')
        if header_norm in _NAGLOWKI_NORM:
            col_indices[_NAGLOWKI_NORM[header_norm]] = i
        else:
            # Kolumna z datą = kolumna absencji; nagłówek jest datą (np. "01.07.2026")
            d = _try_parse_date(header)
            if d:
                date_cols[i] = d

    if 'nazwisko' not in col_indices or 'imie' not in col_indices:
        return {}, [], [f'Arkusz {sheet_name}: brak kolumn Nazwisko/Imię w nagłówku']

    nazwisko_i = col_indices['nazwisko']
    imie_i = col_indices['imie']

    pracownicy: dict[tuple, dict] = {}
    absencje: list[dict] = []

    # Dane pracowników od wiersza 7 (0-indexed: 6)
    for row in rows[6:]:
        if not row:
            continue
        nazwisko_raw = row[nazwisko_i] if nazwisko_i < len(row) else None
        imie_raw = row[imie_i] if imie_i < len(row) else None
        if not nazwisko_raw or not imie_raw:
            continue
        nazwisko = _safe_str(nazwisko_raw)
        imie = _safe_str(imie_raw)
        if not nazwisko or not imie:
            continue

        key = (nazwisko, imie)
        p_data: dict = {'nazwisko': nazwisko, 'imie': imie}
        # Wypełnij pola pracownika ze zmapowanych kolumn
        for field, ci in col_indices.items():
            if ci < len(row) and row[ci] is not None:
                raw = row[ci]
                if field == 'data_zatrudnienia':
                    p_data[field] = _parse_date(raw)
                else:
                    p_data[field] = _safe_str(raw)
        # Jeśli brak działu w kolumnie, użyj działu z nazwy arkusza
        if not p_data.get('dzial'):
            p_data['dzial'] = ARKUSZE_DZIALY.get(sheet_name, '')
        # Zapisz nazwę arkusza — używana jako pole "sektor" w wynikach przydziału
        p_data['_sheet'] = sheet_name
        pracownicy[key] = p_data

        # Sprawdź kolumny dat — jeśli wartość komórki to typ absencji → zapisz nieobecność
        for ci, data_iso in date_cols.items():
            if ci >= len(row):
                continue
            cell = row[ci]
            if cell is None:
                continue
            cell_str = _safe_str(cell)
            # Tylko wartości ze słownika typów absencji są traktowane jako nieobecność
            if cell_str in typy_absencji:
                absencje.append({
                    'pracownik_klucz': f'{nazwisko}|{imie}',
                    'data': data_iso,
                    'typ': cell_str,
                })

    return pracownicy, absencje, []
