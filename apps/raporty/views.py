from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
# apps.rekruci, apps.przydzialy, apps.scoring — moduły legacy (URL-e nie podłączone),
# ale modele i silnik scoringowy są nadal używane przez obsada_excel poniżej
from apps.rekruci.models import Rekrut
from apps.stanowiska.models import Stanowisko
from apps.przydzialy.models import Przydzia
from apps.scoring.engine import ScoringEngine


# Eksportuje raport obsady stanowisk do pliku Excel z dwoma arkuszami
@login_required
def obsada_excel(request):
    wb = openpyxl.Workbook()

    # Styl nagłówka — biały tekst na granatowym tle
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F5C99')

    # Arkusz 1: obsada każdego aktywnego stanowiska
    ws1 = wb.active
    ws1.title = 'Obsada stanowisk'
    headers1 = ['Stanowisko', 'Max pracowników', 'Aktualnie', 'Wolne miejsca', '% zapełnienia', 'Lista pracowników']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    stanowiska = Stanowisko.objects.filter(aktywne=True)
    for s in stanowiska:
        aktywne_przydzialy = Przydzia.objects.filter(stanowisko=s, aktywny=True).select_related('rekrut')
        aktualnie = aktywne_przydzialy.count()
        wolne = s.max_pracownikow - aktualnie
        proc = round(aktualnie / s.max_pracownikow * 100) if s.max_pracownikow else 0
        # Imiona i nazwiska pracowników oddzielone przecinkami w jednej komórce
        pracownicy = ', '.join(str(p.rekrut) for p in aktywne_przydzialy)
        ws1.append([s.nazwa, s.max_pracownikow, aktualnie, wolne, f'{proc}%', pracownicy])

    # auto_size=True jest wskazówką dla openpyxl, ale nie gwarantuje obliczenia width —
    # dlatego fallback: jeśli width jest None (auto_size nie zadziałał), używamy 15
    for col in range(1, 7):
        ws1.column_dimensions[get_column_letter(col)].auto_size = True
        ws1.column_dimensions[get_column_letter(col)].width = max(15, ws1.column_dimensions[get_column_letter(col)].width or 15)

    # Arkusz 2: aktywni rekruci bez przydziału z rekomendowanym stanowiskiem AI
    ws2 = wb.create_sheet('Bez przydziału')
    headers2 = ['Nazwisko', 'Imię', 'Wiek', 'Rekomendowane stanowisko', 'Score AI']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    engine = ScoringEngine()
    rekruci = Rekrut.objects.filter(aktywny=True)
    for r in rekruci:
        # Pomijamy rekrutów, którzy mają już aktywny przydział
        if not r.przydzia_set.filter(aktywny=True).exists():
            wyniki = engine.score(r)
            # Pierwsze stanowisko bez blokad = najlepsze dopasowanie
            top = next((w for w in wyniki if not w['blokady']), None)
            # ws2.append() automatycznie przechodzi do kolejnego wiersza — nie potrzeba licznika row
            ws2.append([
                r.nazwisko, r.imie, r.wiek,
                top['stanowisko'].nazwa if top else 'Brak',
                top['score'] if top else 0,
            ])

    # Wyślij plik Excel jako attachment — przeglądarka pobiera go na dysk
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="obsada_{timezone.now():%Y%m%d}.xlsx"'
    wb.save(response)
    return response
